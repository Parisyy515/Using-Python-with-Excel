from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active

for i in range(1, 20):
    ws.append(range(300))
# fill the worksheet with int value

ws.merge_cells('A1:B5')
ws.unmerge_cells('A1:B5')

ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)
# any merged cell is addressed as their top-left cell, above merge is addressed as B2

cell = ws['B2']
# assign the whole merged cell as cell

cell.font = Font(color=colors.GREEN, size=20, italic=True)
cell.value = 'Merged cell'
cell.alignment = Alignment(horizontal='left', vertical='top')
cell.fill = GradientFill(stop=("000000", "FFFFFF"))
# nice transition use grandient fill, solid color use pattern

wb.save('test.xlsx')
