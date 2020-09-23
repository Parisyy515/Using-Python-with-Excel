from openpyxl.workbook import Workbook
from openpyxl import load_workbook
# import workbook class

wb = Workbook()
# initiat a new workbook for us to work with
ws = wb.active
# create a worksheet aactive under the new workbook

ws1 = wb.create_sheet('Newsheet')
# use create sheet function on workbook
ws2 = wb.create_sheet('Another', 0)
# create a new sheet and index as 0

ws.title = 'Mysheet'
#  altering the tile of current active worksheet

print(wb.sheetnames)
# print out all current worksheet's name, openpyxl wont actually create cells on your worksheet as object until they are accessed

wb2 = load_workbook('regions.xlsx')
# load values into wb2 variable

new_sheet = wb2.create_sheet('NewSheet')
# create a new worksheet in wb2
active_sheet = wb2.active
# grab the active worksheet use active function

cell = active_sheet['A3']
# create a cell variable and assign it with A3 cell position
print(cell.value)

active_sheet['A3'] = 10
# assign a value to a specific cell

wb2.save('updated.xlsx')
# save the value to xlsx
