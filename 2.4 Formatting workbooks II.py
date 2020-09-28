from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active

for i in range(1, 20):
    ws.append(range(300))

ws.merge_cells('A1:B5')
ws.unmerge_cells('A1:B5')

ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)
cell = ws['B2']
# assign the whole merged cell as cell

cell.font = Font(color=colors.GREEN, size=20, italic=True)
cell.value = 'Merged cell'
cell.alignment = Alignment(horizontal='left', vertical='top')
cell.fill = GradientFill(stop=("000000", "FFFFFF"))

wb.save('test.xlsx')

# Named style are objects we can create that store a style so that we can use it multiple times
highlight = NamedStyle(name='highlight')
highlight.font = Font(bold=True)
bd = Side(style='thick', color='000000')

# create a variable to hold the style of side (thick style with black color), fill the cell with solid yellow color
highlight.border = Border(left=bd, top=bd, bottom=bd)
highlight.fill = PatternFill('solid', fgColor='FFFF00')

count = 0
for col in ws.iter_cols(min_col=8, min_row=1, max_col=30, max_row=30):
    col[count].style = highlight
    count = count+1
# in the iter_col function, column is stored as list

wb.save('highlighted.xlsx')
