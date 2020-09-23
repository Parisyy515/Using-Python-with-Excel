from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook('regions.xlsx')
ws = wb.active

c = ws['B6']
# print(c.value)
# access a cell in excel and print its value out

ws['B12'] = 12
# assign the value 12 to one cell

cell_range = ws['A1':'C1']
# print(cell_range)
# accessing many cells through the worksheet by slicing

col_c = ws['C']
# print(col_c)
# index through a specfic column, this can be seen as an array C5 would be col_c[4]

col_range = ws['A':'C']
# index through a block of cell, first index is column, second index is row

row_range = ws[1:5]
# index through a block of cell, first index is row, second index is column

for row in ws.values:
    for value in row:
        print(value)
# iterates over all rows in a worksheet but return just the cell values

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)
# iterates over all cell position in a selected block of excel, iterate row1 then row2

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    for cell in row:
        print(cell)
# iterates over all cell values in a selected block of excel, iterate row1 then row2 with the value flag set to true
